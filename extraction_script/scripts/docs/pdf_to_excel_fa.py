#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pdf_to_excel_fa.py
===================

تبدیل فایل PDF فارسی (شامل صفحات متنی و صفحات جدولی/اکسل‌مانند) به یک
فایل اکسل (.xlsx) با حفظ صحیح ترتیب حروف و کلمات فارسی.

چرا لازم است؟
--------------
وقتی یک PDF فارسی با ابزارهایی مثل pdfplumber / pypdf استخراج می‌شود،
معمولاً حروف هر کلمه و ترتیب کلمات هر خط برعکس می‌شود (چون این کتابخانه‌ها
متن را بر اساس مختصات چپ‌به‌راست می‌خوانند، در حالی که فارسی راست‌به‌چپ
است) و حروف هم اغلب به صورت «شکل ارائه‌ای» (Presentation Form) ذخیره
شده‌اند. این اسکریپت هر دو مشکل را رفع می‌کند:

1. حروف را با Unicode NFKC از حالت Presentation Form به حالت استاندارد
   برمی‌گرداند.
2. ترتیب کلمات را در هر خط، و ترتیب حروف را در هر کلمهٔ فارسی، معکوس
   می‌کند تا متن دوباره خوانا شود (اعداد و کلمات لاتین دست‌نخورده
   می‌مانند).

سپس صفحات را به دو دسته تشخیص می‌دهد:
  - صفحات جدولی (خط‌کشی‌شده، شبیه اکسل) -> به صورت جدول در یک شیت اکسل
  - صفحات متنی (نامه، بخشنامه و ...) -> به صورت خط‌به‌خط در یک شیت اکسل

نصب پیش‌نیازها (در صورت نیاز):
    pip install pdfplumber openpyxl --break-system-packages

نحوهٔ استفاده:
    python pdf_to_excel_fa.py INPUT.pdf [OUTPUT.xlsx]

اگر OUTPUT مشخص نشود، در همان مسیر INPUT با پسوند .xlsx ساخته می‌شود.
"""

import re
import sys
import unicodedata
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------
# ۱) اصلاح ترتیب و شکل متن فارسی
# ----------------------------------------------------------------------

# محدودهٔ یونیکد حروف عربی/فارسی (شامل شکل‌های ارائه‌ای)
_ARABIC_RANGE = re.compile(r"[\u0600-\u06FF\uFB50-\uFDFF\uFE70-\uFEFF]")


def _fix_word(word: str) -> str:
    """اگر کلمه حاوی حروف فارسی/عربی باشد، حروفش را معکوس و نرمال می‌کند.
    اعداد و کلمات انگلیسی بدون تغییر باقی می‌مانند."""
    if _ARABIC_RANGE.search(word):
        return unicodedata.normalize("NFKC", word[::-1])
    return word


def fix_line(line: str) -> str:
    """یک خط را اصلاح می‌کند: ترتیب کلمات را برمی‌گرداند و هر کلمهٔ
    فارسی را با _fix_word اصلاح می‌کند. اعداد/لاتین سر جای منطقی خود
    (نسبت به بقیهٔ خط) باقی می‌مانند."""
    if not line:
        return line
    words = line.split(" ")
    fixed = [_fix_word(w) for w in words]
    return " ".join(reversed(fixed))


def fix_cell(text):
    """برای متن داخل یک سلول جدول که می‌تواند چندخطی باشد."""
    if text is None:
        return None
    lines = text.split("\n")
    return "\n".join(fix_line(l) for l in lines)


# ----------------------------------------------------------------------
# ۲) تبدیل رشته‌های عددی فارسی/انگلیسی به عدد واقعی (برای اکسل)
# ----------------------------------------------------------------------

_PERSIAN_DIGITS = "۰۱۲۳۴۵۶۷۸۹"
_NUMBER_RE = re.compile(r"^-?\d{1,3}(,\d{3})*(\.\d+)?$|^-?\d+(\.\d+)?$")


def to_number(value):
    """اگر مقدار سلول یک عدد باشد (با یا بدون کاما)، به int/float
    تبدیلش می‌کند تا در اکسل قابل جمع و محاسبه باشد؛ در غیر این صورت
    همان رشته را برمی‌گرداند."""
    if value is None:
        return None
    s = value.strip()
    if not s:
        return None
    # ارقام فارسی به انگلیسی
    s2 = s.translate(str.maketrans(_PERSIAN_DIGITS, "0123456789"))
    if _NUMBER_RE.match(s2):
        s2 = s2.replace(",", "")
        try:
            if "." in s2:
                return float(s2)
            return int(s2)
        except ValueError:
            return value
    return value


# ----------------------------------------------------------------------
# ۳) استخراج متن آزاد یک صفحه با ترتیب صحیح خطوط
# ----------------------------------------------------------------------

def extract_ordered_lines(page, top_limit=None):
    """کلمات صفحه را استخراج می‌کند، بر اساس خط (top) گروه‌بندی می‌کند،
    هر خط را از راست به چپ مرتب می‌کند و اصلاح می‌کند.
    اگر top_limit داده شود فقط کلماتی که بالاتر از آن مقدار هستند
    برگردانده می‌شوند (برای گرفتن متن بالای یک جدول)."""
    words = page.extract_words(use_text_flow=False, keep_blank_chars=False, x_tolerance=2)
    # برخی PDFها متن پنهان/خارج از صفحه دارند (کاراکترهاي باقی‌مانده از
    # ویرایش‌هاي قبلی)؛ کلماتی که کاملاً بیرون از کادر صفحه هستند حذف می‌شوند
    words = [
        w for w in words
        if w["x1"] > 0 and w["x0"] < page.width and w["bottom"] > 0 and w["top"] < page.height
    ]
    if top_limit is not None:
        words = [w for w in words if w["top"] < top_limit - 1]
    if not words:
        return []

    # گروه‌بندی خطوط بر اساس نزدیکی مقدار top (برای رفع خطاهای زیرپیکسلی)
    words.sort(key=lambda w: w["top"])
    lines = []
    current_top = None
    current_words = []
    for w in words:
        if current_top is None or abs(w["top"] - current_top) <= 3:
            current_words.append(w)
            current_top = w["top"] if current_top is None else current_top
        else:
            lines.append(current_words)
            current_words = [w]
            current_top = w["top"]
    if current_words:
        lines.append(current_words)

    ordered_lines = []
    for lw in lines:
        lw_sorted = sorted(lw, key=lambda w: -w["x0"])
        text = " ".join(_fix_word(w["text"]) for w in lw_sorted)
        ordered_lines.append(text)
    return ordered_lines


# ----------------------------------------------------------------------
# ۴) استایل شیت (راست‌به‌چپ، فونت، تراز)
# ----------------------------------------------------------------------

FA_FONT = "Tahoma"


def setup_rtl_sheet(ws):
    ws.sheet_view.rightToLeft = True


def style_cell(cell, bold=False, wrap=True, size=11):
    cell.font = Font(name=FA_FONT, size=size, bold=bold)
    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=wrap)


def compute_table_merges(table):
    """محدودهٔ سلول‌های ادغام‌شدهٔ (merged) یک جدول pdfplumber را از روی
    مختصات واقعی سلول‌ها (نه صرفاً خالی/None بودن متن) بازسازی می‌کند.

    نکتهٔ مهم: در جدول‌های واقعی، خیلی از سلول‌ها صرفاً «خالی» هستند
    (بدون متن ولی با کادر جدا)، نه «ادغام‌شده». اتکا به None بودن متن
    برای تشخیص ادغام، سلول‌های خالیِ پشت‌سرهم را به اشتباه در یک سلول
    غول‌پیکر ادغام می‌کند. برای همین از مختصات واقعی هر سلول (که
    pdfplumber بر اساس خطوط جدول محاسبه کرده) استفاده می‌شود: اگر
    عرض/ارتفاع یک سلول از یک ردیف/ستون بیشتر باشد، یعنی واقعاً چند
    ردیف/ستون را در بر گرفته است."""
    rows = table.rows
    n_rows = len(rows)
    if n_rows == 0:
        return []

    xs = set()
    for row in rows:
        for c in row.cells:
            if c:
                xs.add(round(c[0], 1))
                xs.add(round(c[2], 1))
    col_edges = sorted(xs)
    row_tops = [round(r.bbox[1], 1) for r in rows]

    def col_index(x):
        return min(range(len(col_edges)), key=lambda k: abs(col_edges[k] - x))

    def row_index_for_bottom(bottom, start_i):
        r2 = start_i
        for k in range(start_i + 1, n_rows):
            if row_tops[k] < bottom - 0.3:
                r2 = k
            else:
                break
        return r2

    merges = []
    for i, row in enumerate(rows):
        for j, cell in enumerate(row.cells):
            if cell is None:
                continue
            x0, top, x1, bottom = cell
            c1 = col_index(x0)
            c2 = max(c1, col_index(x1) - 1)
            r2 = row_index_for_bottom(round(bottom, 1), i)
            if c2 > j or r2 > i:
                merges.append((i, j, r2, c2))
    return merges


def guess_header_row_count(grid):
    """ردیف‌هایی که اکثر سلول‌های غیرخالی‌شان عدد نیستند را به عنوان
    هدر در نظر می‌گیرد (تا اولین ردیفی که عمدتاً عددی است، یعنی اولین
    ردیف داده)."""
    for i, row in enumerate(grid):
        non_empty = [c for c in row if c not in (None, "")]
        if not non_empty:
            continue
        numeric = [c for c in non_empty if to_number(c) is not None and not isinstance(to_number(c), str)]
        if len(numeric) / len(non_empty) > 0.5:
            return i
    return 1


def autosize_columns(ws, max_width=60):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = max(len(str(l)) for l in str(cell.value).split("\n"))
            col = cell.column_letter
            widths[col] = max(widths.get(col, 10), min(length + 2, max_width))
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# ----------------------------------------------------------------------
# ۵) پردازش هر صفحه
# ----------------------------------------------------------------------

def safe_sheet_name(name: str, used: set) -> str:
    name = re.sub(r'[\\/*?:\[\]]', "", name).strip() or "Sheet"
    name = name[:31]
    base = name
    i = 2
    while name in used:
        suffix = f"_{i}"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def guess_title(lines, fallback):
    for l in lines:
        l = l.strip()
        if l and not l.isdigit():
            return l
    return fallback


def process_page(page, page_no, wb, used_names):
    tables = page.find_tables()
    # فقط جدول‌های واقعی (بیش از ۱ ردیف و ۱ ستون) در نظر گرفته می‌شوند
    real_tables = [t for t in tables if len(t.rows) > 1 and len(t.rows[0].cells) > 1]

    if real_tables:
        for t_idx, table in enumerate(real_tables):
            top_of_table = table.bbox[1]
            header_lines = extract_ordered_lines(page, top_limit=top_of_table) if t_idx == 0 else []
            data = table.extract()
            fixed_rows = [[fix_cell(c) for c in row] for row in data]

            title = guess_title(header_lines, f"صفحه {page_no}")
            sheet_name = safe_sheet_name(
                f"{title[:20]}" if len(real_tables) == 1 else f"{title[:15]}_{t_idx+1}",
                used_names,
            )
            ws = wb.create_sheet(sheet_name)
            setup_rtl_sheet(ws)

            r = 1
            for hl in header_lines:
                cell = ws.cell(row=r, column=1, value=hl)
                style_cell(cell, bold=True, size=12)
                r += 1
            if header_lines:
                r += 1  # یک ردیف خالی فاصله

            table_start_row = r
            merges = compute_table_merges(table)
            header_row_count = guess_header_row_count(fixed_rows)
            n_cols = max(len(row) for row in fixed_rows)

            for row_i, row in enumerate(fixed_rows):
                is_header_row = row_i < header_row_count
                for col_i in range(n_cols):
                    val = row[col_i] if col_i < len(row) else None
                    if val is None:
                        continue  # سلول‌هایی که بخشی از یک ادغام هستند، جدا پر نمی‌شوند
                    write_val = val if is_header_row else to_number(val)
                    cell = ws.cell(row=table_start_row + row_i, column=col_i + 1, value=write_val)
                    style_cell(cell, bold=is_header_row, wrap=True)
                    if is_header_row:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # اعمال ادغام واقعی سلول‌ها، دقیقاً مطابق جدول اصلی PDF
            for r1, c1, r2, c2 in merges:
                ws.merge_cells(
                    start_row=table_start_row + r1, start_column=c1 + 1,
                    end_row=table_start_row + r2, end_column=c2 + 1,
                )

            r = table_start_row + len(fixed_rows)
            autosize_columns(ws)
            ws.freeze_panes = ws.cell(row=table_start_row + header_row_count, column=1)
        return

    # صفحهٔ متنی (بدون جدول)
    lines = extract_ordered_lines(page)
    if not lines:
        return
    title = guess_title(lines, f"صفحه {page_no}")
    sheet_name = safe_sheet_name(title[:25] or f"صفحه_{page_no}", used_names)
    ws = wb.create_sheet(sheet_name)
    setup_rtl_sheet(ws)
    for i, line in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        style_cell(cell, bold=(i == 1), wrap=True, size=11)
    ws.column_dimensions["A"].width = 100


# ----------------------------------------------------------------------
# ۶) تابع اصلی
# ----------------------------------------------------------------------

def convert(pdf_path: str, xlsx_path: str = None) -> str:
    pdf_path = Path(pdf_path)
    if xlsx_path is None:
        xlsx_path = pdf_path.with_suffix(".xlsx")
    else:
        xlsx_path = Path(xlsx_path)

    wb = Workbook()
    wb.remove(wb.active)  # شیت پیش‌فرض خالی را حذف می‌کنیم
    used_names = set()

    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            process_page(page, i, wb, used_names)

    if not wb.sheetnames:
        ws = wb.create_sheet("Sheet1")
        ws["A1"] = "هیچ محتوایی از PDF استخراج نشد."

    wb.save(xlsx_path)
    return str(xlsx_path)


# if __name__ == "__main__":
#     if len(sys.argv) < 2:
#         print("استفاده: python pdf_to_excel_fa.py INPUT.pdf [OUTPUT.xlsx]")
#         sys.exit(1)
#     inp = sys.argv[1]
#     outp = sys.argv[2] if len(sys.argv) > 2 else None
#     out_path = convert(inp, outp)
#     print(f"ذخیره شد: {out_path}")