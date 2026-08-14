from typing import List, Dict, Any, Tuple, Union, Set, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import re
import time
import concurrent.futures


# ============================================================================
# 1. LOW-LEVEL EXCEL LOADING (unchanged)
# ============================================================================

def fill_merged_cells(sheet: Worksheet) -> None:
    merged_ranges = list(sheet.merged_cells.ranges)
    ranges_to_fill: List[Tuple[int, int, int, int, Any]] = []

    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = (
            merged_range.min_col,
            merged_range.min_row,
            merged_range.max_col,
            merged_range.max_row,
        )
        top_left_value = sheet.cell(row=min_row, column=min_col).value
        ranges_to_fill.append((min_col, min_row, max_col, max_row, top_left_value))

    for min_col, min_row, max_col, max_row, _ in ranges_to_fill:
        sheet.unmerge_cells(
            start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col
        )

    for min_col, min_row, max_col, max_row, val in ranges_to_fill:
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row=row, column=col).value = val


def load_excel_without_merges(file_path: str, sheet_name: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    fill_merged_cells(sheet)
    data = list(sheet.values)
    df = pd.DataFrame(data)
    wb.close()
    return df


def get_all_sheet_names(file_path: str) -> List[str]:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    return sheet_names


def load_single_sheet_parallel(file_path: str, sheet_name: str) -> pd.DataFrame:
    return load_excel_without_merges(file_path, sheet_name)


def load_all_sheets_to_memory(file_path: str) -> Dict[str, pd.DataFrame]:
    sheet_names = get_all_sheet_names(file_path)
    loaded_sheets: Dict[str, pd.DataFrame] = {}

    with concurrent.futures.ProcessPoolExecutor() as executor:
        futures = {
            executor.submit(load_single_sheet_parallel, file_path, sheet): sheet
            for sheet in sheet_names
        }
        for future in concurrent.futures.as_completed(futures):
            sheet_name = futures[future]
            try:
                loaded_sheets[sheet_name] = future.result()
            except Exception as e:
                print(f"Error loading sheet '{sheet_name}' in parallel: {e}")

    return loaded_sheets


# ============================================================================
# 2. TEXT HELPERS (unchanged)
# ============================================================================

def normalize_persian_text(text: str) -> str:
    if not isinstance(text, str):
        return ""
    text = text.replace("ي", "ی").replace("ك", "ک").replace("\n", " ").replace("\r", "")
    return re.sub(r"\s+", " ", text).strip()


def normalize_sheet_name(name: str) -> str:
    cleaned = name.replace("\n", "").replace("\r", "").strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned


# ============================================================================
# 3. AUTOMATIC LAYOUT DETECTION  (NEW - replaces hand-written per-sheet config)
# ============================================================================
#
# Every sheet in these budget workbooks follows the same *logical* structure,
# even though the exact row/column boundaries move around from year to year
# or from file to file:
#
#   [ form-header metadata block ]   <- org name / device code / form number / year
#   [ one or more column-header rows ]
#   [ numeric data rows ]
#   [ optional free-text "توضیحات" (notes) block ]
#   [ optional signature footer ("رئیس ...") ]
#
# Instead of hard-coding row numbers per sheet, we classify every row using a
# handful of *content-based* signals (does it look like metadata? mostly
# numbers? does it mention "توضیحات" or "رئیس"?) and derive the boundaries
# from that. This keeps working even if a new export inserts/removes rows,
# reorders forms, or ships forms we haven't seen before - as long as each
# sheet still follows the same general shape.

NUMERIC_PLACEHOLDERS = {"-", "–", "—", "ـ", ""}

# NOTE: these intentionally mirror header_fill()'s regexes (colon / trailing
# digits required) so a column label like "شماره ردیف دستگاه ابلاغ دهنده"
# doesn't get mistaken for the real "کد دستگاه: 12345" metadata line.
FORM_HEADER_PATTERNS = [
    r"(?:عنوان|نام)\s*دستگاه\s*[:-]",
    r"(?:شماره\s*ردیف|کد)\s*دستگاه\s*[:-]?\s*\d+",
    r"فرم\s*شماره\s*\d+",
    r"تفصیلی\s*\d{4}",
    r"اصلاحی[هة]\s*بودجه",
]
DESCRIPTION_PATTERN = r"توضیح"
SIGNATURE_PATTERN = r"رئیس|امضا"


def is_numeric_like(value: Any) -> bool:
    """True for real numbers, and for the '-' placeholder these forms use for zero/empty."""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        s = value.strip()
        if s in NUMERIC_PLACEHOLDERS:
            return True
        try:
            float(s.replace(",", ""))
            return True
        except ValueError:
            return False
    return False


def analyze_row(df: pd.DataFrame, r: int) -> Dict[str, Any]:
    row = df.iloc[r]

    # Collapse consecutive equal values. fill_merged_cells() repeats one value
    # across every cell that used to belong to a merged range, which otherwise
    # inflates the numeric ratio (e.g. a single merged "1397" year label
    # spanning 4 columns looks like 4 numbers instead of 1).
    values: List[Any] = []
    prev_marker: Any = object()
    for v in row:
        if pd.isna(v) or str(v).strip() == "":
            prev_marker = object()
            continue
        if v == prev_marker:
            continue
        values.append(v)
        prev_marker = v

    if not values:
        return {"empty": True, "numeric_ratio": 0.0, "text": ""}

    numeric_count = sum(1 for v in values if is_numeric_like(v))
    combined_text = normalize_persian_text(" ".join(str(v) for v in values))

    return {
        "empty": False,
        "numeric_ratio": numeric_count / len(values),
        "text": combined_text,
        "is_form_header": any(re.search(p, combined_text) for p in FORM_HEADER_PATTERNS),
        "is_description": bool(re.search(DESCRIPTION_PATTERN, combined_text)),
        "is_signature": bool(re.search(SIGNATURE_PATTERN, combined_text)),
    }


def detect_layout(
    df: pd.DataFrame, numeric_threshold: float = 0.5, max_top_scan: int = 10
) -> Dict[str, Any]:
    """
    Scans a raw (merge-filled) sheet DataFrame and infers:
      - form_header:    (start, end) row range holding org/device/form metadata
      - header_rows:    list of row indices to use as (multi-row) column headers
      - data_row:       (start, end) row range holding the numeric data table
      - discription_row:(start, end) row range holding the free-text notes
      - tabular:        False if the sheet has no numeric data table at all
                         (e.g. a pure free-text form like "فرم 3")
    """
    n_rows = df.shape[0]
    rows_meta = [analyze_row(df, r) for r in range(n_rows)]

    # --- form header block: last matching row within the top of the sheet ---
    form_header_end = 0
    for r in range(min(max_top_scan, n_rows)):
        if rows_meta[r].get("is_form_header"):
            form_header_end = r + 1
    if form_header_end == 0:
        form_header_end = 1  # always treat at least the very first row as metadata

    # --- walk down until we hit a numeric-heavy (data) row ---
    data_start: Optional[int] = None
    for r in range(form_header_end, n_rows):
        meta = rows_meta[r]
        if meta["empty"]:
            continue
        if meta.get("is_signature") or meta.get("is_description"):
            break  # hit the footer before any data -> no data table on this sheet
        if meta["numeric_ratio"] >= numeric_threshold:
            data_start = r
            break

    if data_start is None:
        # No tabular block detected (e.g. a narrative/notes-only sheet). Treat
        # everything after the form header as free-text "description" content.
        return {
            "tabular": False,
            "form_header": (0, form_header_end),
            "header_rows": [],
            "data_row": (form_header_end, form_header_end),
            "discription_row": (form_header_end, n_rows),
        }

    header_rows = [r for r in range(form_header_end, data_start) if not rows_meta[r]["empty"]]

    # --- find where the data table ends: a "توضیحات" marker and/or a signature line ---
    desc_row: Optional[int] = None
    sign_row: Optional[int] = None
    for r in range(data_start, n_rows):
        meta = rows_meta[r]
        if meta["empty"]:
            continue
        if desc_row is None and meta.get("is_description"):
            desc_row = r
        # A row that mentions a signatory title (e.g. "رئیس ...") *inside* a
        # description paragraph is still descriptive content, not the actual
        # signature footer - only treat it as the footer once the row is no
        # longer also flagged as description text.
        if sign_row is None and meta.get("is_signature") and not meta.get("is_description"):
            sign_row = r
            break

    if desc_row is not None:
        data_end = desc_row
        discription_end = sign_row if sign_row is not None else n_rows
        discription_row = (desc_row, discription_end)
    else:
        data_end = sign_row if sign_row is not None else n_rows
        discription_row = (data_end, data_end)  # empty block, nothing to report

    return {
        "tabular": True,
        "form_header": (0, form_header_end),
        "header_rows": header_rows,
        "data_row": (data_start, data_end),
        "discription_row": discription_row,
    }


def detect_hierarchy_columns(
    df: pd.DataFrame, data_row_range: Tuple[int, int], text_threshold: float = 0.6
) -> List[int]:
    """
    Finds the column(s) that hold row *labels* (e.g. "عنوان") rather than
    budget figures, by checking which columns are mostly non-numeric text
    within the data-row range. These are the columns process_sheet() uses to
    build the row index (hierarchy_cols_indices) and then drops from the
    final value table.
    """
    start, end = data_row_range
    sub = df.iloc[start:] if end == -1 else df.iloc[start:end]
    n_cols = df.shape[1]

    hierarchy_cols: List[int] = []
    scores: List[Tuple[int, float, int]] = []
    for c in range(n_cols):
        col = sub.iloc[:, c]
        non_null = [v for v in col if pd.notna(v) and str(v).strip() != ""]
        if not non_null:
            scores.append((c, 0.0, 0))
            continue
        text_count = sum(1 for v in non_null if not is_numeric_like(v))
        text_ratio = text_count / len(non_null)
        scores.append((c, text_ratio, len(non_null)))
        if text_ratio >= text_threshold:
            hierarchy_cols.append(c)

    if not hierarchy_cols:
        # Fallback: nothing crossed the threshold (unusual sheet) - just use
        # whichever column looks "most label-like" so we still have an index.
        best = max(scores, key=lambda t: t[1] * t[2], default=None)
        if best and best[1] > 0:
            hierarchy_cols = [best[0]]

    return hierarchy_cols


# ============================================================================
# 4. METADATA EXTRACTION (unchanged)
# ============================================================================

def form_metadata(df: pd.DataFrame, metadata_start_row: int, metadata_end_row: int) -> Set[str]:
    if metadata_end_row == -1:
        metadata_df = df.iloc[metadata_start_row:]
    else:
        metadata_df = df.iloc[metadata_start_row:metadata_end_row]

    metadata: Set[str] = set()
    for _, row in metadata_df.iterrows():
        row_cells = [
            str(val).strip() for val in row if pd.notna(val) and str(val).strip() != ""
        ]
        for cell_text in row_cells:
            metadata.add(cell_text)
    return metadata


def header_fill(data_set: Set[str]) -> Dict[str, Any]:
    header: Dict[str, Any] = {
        "org_name": None,
        "device_id": None,
        "budget_year": None,
        "form_number": None,
        "form_title": None,
    }
    valid_cells: Set[str] = set()
    for cell in data_set:
        clean_cell = normalize_persian_text(cell)
        if not clean_cell:
            continue
        if "ریال" in clean_cell:
            continue
        valid_cells.add(clean_cell)

    cells_to_discard: Set[str] = set()

    for cell in valid_cells:
        matched_any = False

        if not header["org_name"]:
            m_org = re.search(r"(?:عنوان|نام)\s*دستگاه\s*[:-]?\s*(.+)", cell)
            if m_org:
                header["org_name"] = m_org.group(1).strip()
                matched_any = True

        if not header["device_id"]:
            m_dev = re.search(r"(?:شماره\s*ردیف|کد)\s*دستگاه\s*[:-]?\s*(\d+)", cell)
            if m_dev:
                header["device_id"] = m_dev.group(1).strip()
                matched_any = True

        if not header["budget_year"]:
            m_year = re.search(r"تفصیلی\s*(\d{4})", cell)
            if m_year:
                header["budget_year"] = int(m_year.group(1))
                matched_any = True

        if not header["form_number"]:
            m_form = re.search(r"فرم\s*شماره\s*(\d+(?:\s*[\-\/]\s*\d+)?)(.*)", cell)
            if m_form:
                raw_form_num = m_form.group(1).strip()
                cleaned_form_num = re.sub(r"\s*([\-\/])\s*", r"\1", raw_form_num)
                header["form_number"] = cleaned_form_num

                potential_title = m_form.group(2).strip(" -_:")
                if potential_title and len(potential_title) > 3:
                    header["form_title"] = potential_title
                matched_any = True

        if matched_any:
            cells_to_discard.add(cell)

    valid_cells -= cells_to_discard

    if not header["form_title"]:
        best_candidate: Optional[str] = None
        for cell in valid_cells:
            if len(cell) > 5 and not re.match(r"^[\d,\.\-\/]+$", cell):
                if best_candidate is None or len(cell) > len(best_candidate):
                    best_candidate = cell
        if best_candidate:
            header["form_title"] = best_candidate
    return header


def reconstruct_headers(df: pd.DataFrame, header_rows: List[int], delimiter: str = " _ ") -> List[str]:
    new_headers: List[str] = []
    num_cols = df.shape[1]

    for col_idx in range(num_cols):
        col_header_values: List[str] = []
        for r in header_rows:
            val = df.iloc[r, col_idx]
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str:
                    col_header_values.append(val_str)

        deduplicated_values: List[str] = []
        for val in col_header_values:
            if not deduplicated_values or val != deduplicated_values[-1]:
                deduplicated_values.append(val)

        if deduplicated_values:
            new_headers.append(delimiter.join(deduplicated_values))
        else:
            new_headers.append(f"Unnamed_{col_idx}")

    return new_headers


# ============================================================================
# 5. MAIN PER-SHEET PROCESSOR
# ============================================================================

_DESCRIPTION_LABELS = {"توضیحات", "توضيحات"}


def process_sheet(
    df: pd.DataFrame,
    header_rows: Optional[List[int]] = None,
    hierarchy_cols_indices: Optional[List[int]] = None,
    data_row: Optional[List[int]] = None,
    discription_row: Optional[List[int]] = None,
    form_header: Optional[List[int]] = None,
    delimiter: str = " _ ",
) -> Optional[Dict[str, Any]]:
    """
    Processes one sheet into a clean, indexed DataFrame + metadata.

    All of header_rows / hierarchy_cols_indices / data_row / discription_row /
    form_header are now OPTIONAL. Pass them explicitly to force specific
    boundaries (e.g. for one-off exceptional sheets); leave them as None (the
    default) to have the layout auto-detected from the sheet's content, which
    is what makes this work across differently-laid-out workbooks.
    """
    layout = None
    if header_rows is None or data_row is None or discription_row is None or form_header is None:
        layout = detect_layout(df)
        if not layout["tabular"]:
            # No numeric data table on this sheet (e.g. a narrative-only form).
            # Just surface whatever metadata/description we can find.
            header = header_fill(form_metadata(df, *layout["form_header"]))
            discription = form_metadata(df, *layout["discription_row"])
            discription = {c for c in discription if normalize_persian_text(c) not in _DESCRIPTION_LABELS}
            return {"data": None, "metadata": discription, "header": header}

        header_rows = header_rows if header_rows is not None else layout["header_rows"]
        data_row = data_row if data_row is not None else list(layout["data_row"])
        discription_row = discription_row if discription_row is not None else list(layout["discription_row"])
        form_header = form_header if form_header is not None else list(layout["form_header"])

    if hierarchy_cols_indices is None:
        hierarchy_cols_indices = detect_hierarchy_columns(df, tuple(data_row))

    # 0. Reconstruct headers
    headers = reconstruct_headers(df, header_rows, delimiter=delimiter)

    # 1. info
    discription = form_metadata(df, discription_row[0], discription_row[1])
    discription = {c for c in discription if normalize_persian_text(c) not in _DESCRIPTION_LABELS}
    header = header_fill(form_metadata(df, form_header[0], form_header[1]))

    # 2. Slice data rows
    if data_row[1] == -1:
        clean_df = df.iloc[data_row[0]:].copy()
    else:
        clean_df = df.iloc[data_row[0]:data_row[1]].copy()
    clean_df.columns = headers

    # 3. Build the combined index
    combined_index: List[str] = []
    for _, row in clean_df.iterrows():
        seen = set()
        unique_parts: List[str] = []
        for idx in hierarchy_cols_indices:
            val = row.iloc[idx]
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str and val_str not in seen:
                    seen.add(val_str)
                    unique_parts.append(val_str)
        combined_index.append(delimiter.join(unique_parts))

    clean_df.index = combined_index
    clean_df = clean_df[clean_df.index.str.strip() != ""]

    # 4. Safely drop the hierarchy columns using their integer positions
    cols_to_keep = [col for i, col in enumerate(clean_df.columns) if i not in hierarchy_cols_indices]
    clean_df = clean_df[cols_to_keep]

    unnamed_cols = [
        col for col in clean_df.columns
        if str(col) == "Unnamed" or str(col).startswith("Unnamed_") or str(col).startswith("Unnamed:")
    ]
    clean_df.drop(columns=unnamed_cols, inplace=True, errors="ignore")

    return {"data": clean_df, "metadata": discription, "header": header}


# ============================================================================
# 6. TOP-LEVEL DRIVER
# ============================================================================

def main(budget_type: str, excel_file_path: Dict[str, str], sheet_to_config_map: Optional[Dict[str, List[str]]] = None,
         forms_param: Optional[Dict[str, Dict[str, Any]]] = None):
    """
    sheet_to_config_map / forms_param are now OPTIONAL. If omitted, every
    sheet in the workbook is auto-detected and processed under its own
    (normalized) sheet name. If provided, they behave as before: a sheet can
    still be manually re-labelled/forced via FORMS_PARAM, but nothing is
    required to be hand-mapped anymore for the auto path to work.
    """
    start_time = time.time()

    if budget_type not in ("تفضیلی", "اصلاحیه", "ابلاغ", "تاییدیه"):
        raise ValueError("Invalid Budget type specified.")

    if not excel_file_path.get(budget_type):
        raise ValueError(f"Excel file path for budget type '{budget_type}' is not provided.")

    raw_sheets_in_ram = load_all_sheets_to_memory(excel_file_path[budget_type])

    result_sheets: Dict[str, Any] = {}
    for original_sheet_name, df_sheet in raw_sheets_in_ram.items():
        normalized_name = normalize_sheet_name(original_sheet_name)

        # Manual override path (kept for backward compatibility / exceptional sheets)
        if sheet_to_config_map and normalized_name in sheet_to_config_map:
            for config_key in sheet_to_config_map[normalized_name]:
                if forms_param and config_key in forms_param:
                    param = forms_param[config_key]
                    print(f"Processing '{original_sheet_name}' mapped as '{config_key}' (manual boundaries)")
                    result_sheets[config_key] = process_sheet(
                        df=df_sheet,
                        header_rows=param.get("header_rows"),
                        data_row=param.get("data_row"),
                        discription_row=param.get("discription_row"),
                        form_header=param.get("form_header"),
                        hierarchy_cols_indices=param.get("hierarchy_cols_indices"),
                    )
                    continue

        # Default path: fully auto-detected
        print(f"Processing '{original_sheet_name}' (auto-detected boundaries)")
        result_sheets[normalized_name] = process_sheet(df=df_sheet)

    end_time = time.time()
    print("runtime elapsed:", end_time - start_time)
    return result_sheets

