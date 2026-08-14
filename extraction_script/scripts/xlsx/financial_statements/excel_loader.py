"""
توابع پایه و مشترک برای خواندن اکسل و پاک‌سازی سلول‌های ادغام‌شده.
این بخش دقیقا همان منطق کد قبلی شماست و بدون تغییر قابل استفاده مجدد است.
"""
from typing import List, Dict, Any, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import re
import concurrent.futures


def fill_merged_cells(sheet: Worksheet) -> None:
    merged_ranges = list(sheet.merged_cells.ranges)
    ranges_to_fill: List[Tuple[int, int, int, int, Any]] = []

    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = (
            merged_range.min_col,
            merged_range.min_row,
            merged_range.max_col,
            merged_range.max_row,
        )
        top_left_value = sheet.cell(row=min_row, column=min_col).value
        ranges_to_fill.append((min_col, min_row, max_col, max_row, top_left_value))

    for min_col, min_row, max_col, max_row, _ in ranges_to_fill:
        sheet.unmerge_cells(
            start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col
        )

    for min_col, min_row, max_col, max_row, val in ranges_to_fill:
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row=row, column=col).value = val


def load_excel_without_merges(file_path: str, sheet_name: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    fill_merged_cells(sheet)
    data = list(sheet.values)
    df = pd.DataFrame(data)
    wb.close()
    return df


def get_all_sheet_names(file_path: str) -> List[str]:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    return sheet_names


def load_single_sheet_parallel(file_path: str, sheet_name: str) -> pd.DataFrame:
    return load_excel_without_merges(file_path, sheet_name)


def load_all_sheets_to_memory(file_path: str, parallel: bool = True) -> Dict[str, pd.DataFrame]:
    """
    هر شیت را با ThreadPoolExecutor به‌صورت موازی بارگذاری می‌کند (نه
    ProcessPoolExecutor)، چون این کار I/O-bound است (خواندن xlsx) و از
    Process-spawn استفاده نمی‌کند؛ این موضوع باعث سازگاری کامل با محیط‌هایی
    مثل Streamlit می‌شود که در آن‌ها ساخت پردازش‌های سیستم‌عاملی جدید
    (ProcessPoolExecutor روی ویندوز) با خطای "process pool terminated
    abruptly" مواجه می‌شود. اگر موازی‌سازی به هر دلیلی شکست بخورد، به حالت
    سریالی برمی‌گردد.
    """
    sheet_names = get_all_sheet_names(file_path)
    loaded_sheets: Dict[str, pd.DataFrame] = {}

    if not parallel or len(sheet_names) <= 1:
        for sheet in sheet_names:
            try:
                loaded_sheets[sheet] = load_excel_without_merges(file_path, sheet)
            except Exception as e:
                print(f"خطا در بارگذاری شیت '{sheet}': {e}")
        return loaded_sheets

    try:
        with concurrent.futures.ThreadPoolExecutor() as executor:
            futures = {
                executor.submit(load_single_sheet_parallel, file_path, sheet): sheet
                for sheet in sheet_names
            }
            for future in concurrent.futures.as_completed(futures):
                sheet_name = futures[future]
                try:
                    loaded_sheets[sheet_name] = future.result()
                except Exception as e:
                    print(f"خطا در بارگذاری شیت '{sheet_name}' به صورت موازی: {e}")
    except Exception as e:
        print(f"موازی‌سازی بارگذاری شیت‌ها شکست خورد، بازگشت به حالت سریالی: {e}")
        loaded_sheets = {}

    # اگر به هر دلیلی (مثلا محیط اجرای غیرمعمول) بارگذاری موازی برخی شیت‌ها را
    # از دست داده باشد، به‌صورت سریالی امتحان مجدد می‌کنیم تا داده‌ای از دست نرود.
    missing_sheets = [s for s in sheet_names if s not in loaded_sheets]
    for sheet in missing_sheets:
        try:
            loaded_sheets[sheet] = load_excel_without_merges(file_path, sheet)
        except Exception as e2:
            print(f"خطا در بارگذاری شیت '{sheet}': {e2}")

    return loaded_sheets


def normalize_persian_text(text: str) -> str:
    if not isinstance(text, str):
        return ""
    text = text.replace("ي", "ی").replace("ك", "ک").replace("\n", " ").replace("\r", "")
    return re.sub(r"\s+", " ", text).strip()


def normalize_sheet_name(name: str) -> str:
    cleaned = name.replace("\n", "").replace("\r", "").strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned
