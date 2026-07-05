from config import EXCEL_FILE_PATH, FORMS_PARAM, SHEET_TO_CONFIG_MAP
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import io
import re
from typing import List, Tuple, Any, Dict
import concurrent.futures
import time
# from IPython.display import display


def fill_merged_cells(sheet: Worksheet) -> None:
    merged_ranges = list(sheet.merged_cells.ranges)
    ranges_to_fill: List[Tuple[int, int, int, int, Any]] = []
    
    # 1. Capture the ranges and their source top-left values
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = (
            merged_range.min_col,
            merged_range.min_row,
            merged_range.max_col,
            merged_range.max_row,
        )
        top_left_value = sheet.cell(row=min_row, column=min_col).value
        ranges_to_fill.append((min_col, min_row, max_col, max_row, top_left_value))
        
    # 2. Unmerge cells to make them writable standard cells
    for min_col, min_row, max_col, max_row, _ in ranges_to_fill:
        sheet.unmerge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col
        )
        
    # 3. Fill all cells within the former merged ranges with the top-left value
    for min_col, min_row, max_col, max_row, val in ranges_to_fill:
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row=row, column=col).value = val


def load_excel_without_merges(file_path: str, sheet_name: str) -> pd.DataFrame:
    # wb = openpyxl.load_workbook(file_path, data_only=True)
    # sheet = wb[sheet_name]
    
    # fill_merged_cells(sheet)
    
    # virtual_file = io.BytesIO()
    # wb.save(virtual_file)
    # virtual_file.seek(0)
    
    # df = pd.read_excel(virtual_file, sheet_name=sheet_name, header=None)
    # return df
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    fill_merged_cells(sheet)
    
    # Blazing fast direct conversion from memory to DataFrame (Bypasses wb.save and disk I/O)
    data = list(sheet.values)
    df = pd.DataFrame(data)
    wb.close()
    return df

def get_all_sheet_names(file_path: str) -> List[str]:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    return sheet_names


def load_single_sheet_parallel(file_path: str, sheet_name: str) -> pd.DataFrame:
    return load_excel_without_merges(file_path, sheet_name)

def load_all_sheets_to_memory(file_path: str) -> Dict[str, pd.DataFrame]:
    sheet_names = get_all_sheet_names(file_path)
    loaded_sheets: Dict[str, pd.DataFrame] = {}
    
    with concurrent.futures.ProcessPoolExecutor() as executor:
        futures = {
            executor.submit(load_single_sheet_parallel, file_path, sheet): sheet 
            for sheet in sheet_names
        }
        
        for future in concurrent.futures.as_completed(futures):
            sheet_name = futures[future]
            try:
                df_cleaned = future.result()
                loaded_sheets[sheet_name] = df_cleaned
            except Exception as e:
                print(f"Error loading sheet '{sheet_name}' in parallel: {e}")
                
    return loaded_sheets



def reconstruct_headers(df: pd.DataFrame, header_rows: List[int], delimiter: str = " _ ") -> List[str]:
    new_headers: List[str] = []
    num_cols = df.shape[1]
    
    for col_idx in range(num_cols):
        col_header_values: List[str] = []
        
        for r in header_rows:
            val = df.iloc[r, col_idx]
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str:
                    col_header_values.append(val_str)
        
        deduplicated_values: List[str] = []
        for val in col_header_values:
            if not deduplicated_values or val != deduplicated_values[-1]:
                deduplicated_values.append(val)
                
        if deduplicated_values:
            new_headers.append(delimiter.join(deduplicated_values))
        else:
            new_headers.append(f"Unnamed_{col_idx}")
            
    return new_headers

def form_metadata(df: pd.DataFrame, metadata_start_row: int, metadata_end_row: int) -> Dict[Set, Any]:
    # if metadata_start_row <= df.shape[0] or metadata_end_row >= df.shape[0]:
    #     return {"metadata": ""}            

    if metadata_end_row == -1:
        metadata_df = df.iloc[metadata_start_row:]
    else : metadata_df = df.iloc[metadata_start_row:metadata_end_row]
        
    metadata: Set[str] = set()

    for _, row in metadata_df.iterrows():
        row_cells = [
            str(val).strip() for val in row 
            if pd.notna(val) and str(val).strip() != ""
        ]
        
        for cell_text in row_cells:
            metadata.add(cell_text)
                
    return metadata


def process_sheet(
    df: pd.DataFrame,
    header_rows: List[int], 
    start_data_row: int, 
    hierarchy_cols_indices: List[int],
    end_data_row: int = -1,
    start_metadata_row: int = -1,
    end_metadata_row: int = -1,
    delimiter: str = " _ "
) -> pd.DataFrame:
    
    # 0. Reconstruct headers
    headers = reconstruct_headers(df, header_rows, delimiter=delimiter)
    
    # 1. metadata
    metadata = form_metadata(df, start_metadata_row, end_metadata_row)

    # 2. Slice data rows
    if end_data_row == -1:
        clean_df = df.iloc[start_data_row:].copy()
    else: clean_df = df.iloc[start_data_row:end_data_row].copy()
    clean_df.columns = headers
    
    # 3. Build the combined index
    combined_index: List[str] = []
    for _, row in clean_df.iterrows():
        seen = set()
        unique_parts: List[str] = []
        for idx in hierarchy_cols_indices:
            val = row.iloc[idx]
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str and val_str not in seen:
                    seen.add(val_str)
                    unique_parts.append(val_str)
        combined_index.append(delimiter.join(unique_parts))
        
    # Assign the generated index
    clean_df.index = combined_index
    clean_df = clean_df[clean_df.index.str.strip() != ""]

    # 4. Safely drop the hierarchy columns using their integer positions
    cols_to_keep = [
        col for i, col in enumerate(clean_df.columns) 
        if i not in hierarchy_cols_indices
    ]
    clean_df = clean_df[cols_to_keep]

    unnamed_cols = [
        col for col in clean_df.columns 
        if str(col) == "Unnamed" or str(col).startswith("Unnamed_") or str(col).startswith("Unnamed:")
    ]
    clean_df.drop(columns=unnamed_cols, inplace=True, errors='ignore')
    return {'data': clean_df, 'metadata': metadata}

def normalize_sheet_name(name: str) -> str:
    cleaned = name.replace("\n", "").replace("\r", "").strip()
    cleaned = re.sub(r'\s+', ' ', cleaned)
    return cleaned



if __name__ == "__main__":
    t1 = time.time()
    raw_sheets_in_ram = load_all_sheets_to_memory(EXCEL_FILE_PATH)
    
    # Process sheets dynamically using the clean mapping
    result_sheets = dict()
    for original_sheet_name, df_sheet in raw_sheets_in_ram.items():
        normalized_name = normalize_sheet_name(original_sheet_name)
        
        if normalized_name in SHEET_TO_CONFIG_MAP:
            
            target_configs = SHEET_TO_CONFIG_MAP[normalized_name]
            
            for config_key in target_configs:
                # if config_key in FORMS_PARAM and config_key in ['form1']:
                if config_key in FORMS_PARAM:
                    param = FORMS_PARAM[config_key]
                    print(f"Processing '{original_sheet_name}' mapped as '{config_key}'")
                    
                    # display(df_sheet)

                    result_sheets[config_key] = process_sheet(
                        df=df_sheet,
                        header_rows=param["header_rows"],
                        start_data_row=param["start_data_row"],
                        end_data_row=param["end_data_row"],
                        start_metadata_row=param["start_metadata_row"],
                        end_metadata_row=param["end_metadata_row"],
                        hierarchy_cols_indices=param["hierarchy_cols_indices"]
                    )
                    # display(result_sheets[config_key]["metadata"])
                    # display(result_sheets[config_key]["data"])
                    # print(result_sheets[config_key].columns)
        else:
            print(f"Skipped: Sheet '{original_sheet_name}' (Normalized: '{normalized_name}') has no mapping defined.")

    result_sheets
    t2 = time.time()
    print("runtime:", t2-t1)

#TODO: document of script

