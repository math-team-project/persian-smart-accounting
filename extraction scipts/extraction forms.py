import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import io
from typing import List, Tuple, Any, Dict


EXCEL_FILE_PATH = "Budget.xlsx"


def fill_merged_cells(sheet: Worksheet) -> None:
    merged_ranges = list(sheet.merged_cells.ranges)
    ranges_to_fill: List[Tuple[int, int, int, int, Any]] = []
    
    # 1. Capture the ranges and their source top-left values
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = (
            merged_range.min_col,
            merged_range.min_row,
            merged_range.max_col,
            merged_range.max_row,
        )
        top_left_value = sheet.cell(row=min_row, column=min_col).value
        ranges_to_fill.append((min_col, min_row, max_col, max_row, top_left_value))
        
    # 2. Unmerge cells to make them writable standard cells
    for min_col, min_row, max_col, max_row, _ in ranges_to_fill:
        sheet.unmerge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col
        )
        
    # 3. Fill all cells within the former merged ranges with the top-left value
    for min_col, min_row, max_col, max_row, val in ranges_to_fill:
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row=row, column=col).value = val


def load_excel_without_merges(file_path: str, sheet_name: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    
    fill_merged_cells(sheet)
    
    virtual_file = io.BytesIO()
    wb.save(virtual_file)
    virtual_file.seek(0)
    
    df = pd.read_excel(virtual_file, sheet_name=sheet_name, header=None)
    return df


def get_all_sheet_names(file_path: str) -> List[str]:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    return sheet_names


def load_all_sheets_to_memory(file_path: str) -> Dict[str, pd.DataFrame]:
    sheet_names = get_all_sheet_names(file_path)
    print("Sheets found in file:", sheet_names)
    
    loaded_sheets: Dict[str, pd.DataFrame] = {}
    
    for sheet in sheet_names:   #TODO: speedup loading the sheets
        try:
            df_cleaned = load_excel_without_merges(file_path, sheet)
            loaded_sheets[sheet] = df_cleaned
        except Exception as e:
            print(f"Error loading sheet '{sheet}': {e}")
            
    return loaded_sheets



def reconstruct_headers(df: pd.DataFrame, header_rows: List[int], delimiter: str = " _ ") -> List[str]:
    new_headers: List[str] = []
    num_cols = df.shape[1]
    
    for col_idx in range(num_cols):
        col_header_values: List[str] = []
        
        for r in header_rows:
            val = df.iloc[r, col_idx]
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str:
                    col_header_values.append(val_str)
        
        deduplicated_values: List[str] = []
        for val in col_header_values:
            if not deduplicated_values or val != deduplicated_values[-1]:
                deduplicated_values.append(val)
                
        if deduplicated_values:
            new_headers.append(delimiter.join(deduplicated_values))
        else:
            new_headers.append(f"Unnamed_{col_idx}")
            
    return new_headers


def process_sheet(
    df: pd.DataFrame,       #TODO: add excel to pandas def
    header_rows: List[int], 
    start_data_row: int, 
    hierarchy_cols_indices: List[int],
    end_data_row: int = -1,
    delimiter: str = " _ "
) -> pd.DataFrame:
    
    # 1. Reconstruct headers
    headers = reconstruct_headers(df, header_rows, delimiter=delimiter)
    
    # 2. Slice data rows
    if end_data_row == -1:
        clean_df = df.iloc[start_data_row:].copy()
    else: clean_df = df.iloc[start_data_row:end_data_row].copy()
    clean_df.columns = headers
    
    # 3. Build the combined index
    combined_index: List[str] = []
    for _, row in clean_df.iterrows():
        seen = set()
        unique_parts: List[str] = []
        for idx in hierarchy_cols_indices:
            val = row.iloc[idx]
            if pd.notna(val):
                val_str = str(val).strip()
                if val_str and val_str not in seen:
                    seen.add(val_str)
                    unique_parts.append(val_str)
        combined_index.append(delimiter.join(unique_parts))
        
    # Assign the generated index
    clean_df.index = combined_index
    clean_df = clean_df[clean_df.index.str.strip() != ""]

    # 4. Safely drop the hierarchy columns using their integer positions
    cols_to_keep = [
        col for i, col in enumerate(clean_df.columns) 
        if i not in hierarchy_cols_indices
    ]
    clean_df = clean_df[cols_to_keep]

    clean_df.drop(columns=["Unnamed"], inplace=True, errors='ignore')
    #TODO: get truly the توضیحات
    return clean_df


if __name__ == "__main__":
    raw_sheets_in_ram = load_all_sheets_to_memory(EXCEL_FILE_PATH)

    FORMS_PARAM = {'form1': {'header_rows':[3, 4], 'start_data_row':5, 'end_data_row':-1, 'hierarchy_cols_indices':[14, 13, 12]},
                   'form2': {'header_rows':[2, 3], 'start_data_row':4, 'end_data_row':-1, 'hierarchy_cols_indices':[14, 13]},
                   'form4': {'header_rows':[6,7,8], 'start_data_row':9, 'end_data_row':-1, 'hierarchy_cols_indices':range(20, 18-1, -1)},
                   'form5': {'header_rows':[3, 4], 'start_data_row':5, 'end_data_row':-1, 'hierarchy_cols_indices':range(14, 11-1, -1)},
                   'form5-1a': {'header_rows':[3, 4], 'start_data_row':5, 'end_data_row':16, 'hierarchy_cols_indices':range(13, 10-1, -1)},
                   'form6a': {'header_rows':[4,5], 'start_data_row':6, 'end_data_row':15, 'hierarchy_cols_indices':range(12, 11-1, -1)},
                   'form6b': {'header_rows':[17,18], 'start_data_row':19, 'end_data_row':-1, 'hierarchy_cols_indices':range(12, 11-1, -1)},
                   'form7': {'header_rows':[4,5], 'start_data_row':6, 'end_data_row':-1, 'hierarchy_cols_indices':range(14, 13-1, -1)},
                   'form8': {'header_rows':[4,5], 'start_data_row':6, 'end_data_row':-1, 'hierarchy_cols_indices':range(10, 7-1, -1)},
                   'form9': {'header_rows':[4,5], 'start_data_row':6, 'end_data_row':-1, 'hierarchy_cols_indices':range(7, 6-1, -1)},
                   'form10': {'header_rows':[4,5], 'start_data_row':6, 'end_data_row':-1, 'hierarchy_cols_indices':[7]},
                   'form5-1b': {'header_rows':[3, 4], 'start_data_row':21, 'end_data_row':27, 'hierarchy_cols_indices':range(13, 10-1, -1)},
                   }
    
    result_sheets = Dict()
    for i, df_sheet in enumerate(raw_sheets_in_ram):
        result_sheets[f'form{i}'] = process_sheet(df=df_sheet, header_rows = FORMS_PARAM[f'form{i}']["header_rows"],
                                                  start_data_row = FORMS_PARAM[f'form{i}']["start_data_row"],
                                                  header_rows = FORMS_PARAM[f'form{i}']["header_rows"],
                                                  hierarchy_cols_indices = FORMS_PARAM[f'form{i}']["hierarchy_cols_indices"])

    result_sheets